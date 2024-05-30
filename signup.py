import tkinter as tk
from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk
import pygame
import openpyxl 
import pandas as pd
import pygame


       
root = tk.Tk()
bg=ImageTk.PhotoImage(file='timetravel.png')
bg_image=Label(root,image=bg).place(x=0,y=0, relwidth=1,relheight=1)
root.state('zoomed')
root.geometry('935x500+300+200')
root.configure(bg='#fff')
root.resizable(False,False)

def save_user_profile():
    user_name = user.get()
    user_id = password.get()
    
    user_profile = openpyxl.load_workbook("userprofile.xlsx")
    profile_sheet = user_profile.active

    profile_sheet[f'A{profile_sheet.max_row+1 }'] = user_name
    profile_sheet[f'B{profile_sheet.max_row }'] = user_id
    
    user_profile.save("userprofile.xlsx")


frame=Frame(root,width=350,height=350,bg='white')
frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

heading=Label(frame,text='Sign Up',fg='#57a1f8',bg='white',font=('Microsoft YahHei UI Light',23,'bold'))
heading.place(x=130, y=5)

#############################################################################################################

def on_enter(e):
    user.delete(0,'end')
    
def on_leave(e):
    name=user.get()
    if name=='':
        user.insert(0, 'Username')

user = Entry(frame,width=25, fg='black',border=0, bg='white',highlightthickness=0,font=('Microsoft YahHei UI Light',11))
user.place(x=30, y=80)
user.insert(0,'Username')
user.bind('<FocusIn>',on_enter)
user.bind('<FocusOut>', on_leave)

Frame(frame,width=295,height=2,bg='black').place(x=25,y=107)

#############################################################################################################

def on_enter(e):
    password.delete(0,'end')
    
def on_leave(e):
    pw=password.get()
    if pw=='':
        password.insert(0, 'Password')

password = Entry(frame,width=25, fg='black',border=0, bg='white',highlightthickness=0,font=('Microsoft YahHei UI Light',11))
password.place(x=30, y=150)
password.insert(0,'Password')
password.bind('<FocusIn>',on_enter)
password.bind('<FocusOut>', on_leave)

Frame(frame,width=295,height=2,bg='black').place(x=25,y=177)

##############################################################################################################

Button(frame,width=20,pady=7,text='Sign in',bg='#57a1f8',fg='white',border=0, command=save_user_profile).place(x=70,y=204)

label_error_name = tk.Label(frame, text="", fg="red", bg="black")
label_error_id = tk.Label(frame, text="", fg="red", bg="black")


root.mainloop()